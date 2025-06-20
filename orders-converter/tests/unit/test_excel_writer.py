import os
import tempfile
import pandas as pd
from openpyxl import load_workbook
from orders_converter.io.excel_writer import write_order_excel

def test_write_order_excel_creates_file_and_sheets():
    df = pd.DataFrame({
        'Item': ['A', 'B'],
        'Qty': [1, 2],
        'Price': [10.5, 20.0]
    })
    meta = {'po_number': '123', 'vendor_number': 'V001', 'ship_by_date': '2024-06-01', 'payment_terms': 'Net 30', 'total': '30.5', 'page_count': 1}
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, 'test.xlsx')
        write_order_excel(df, meta, out_path)
        assert os.path.exists(out_path)
        wb = load_workbook(out_path)
        assert 'Order' in wb.sheetnames
        assert 'Summary' in wb.sheetnames 