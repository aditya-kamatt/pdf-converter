import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers
from openpyxl import load_workbook

def write_order_excel(df: pd.DataFrame, meta: dict, out_path: str):
    """
    Write the order DataFrame and meta info to an Excel file with formatting.
    - df: main table as DataFrame
    - meta: dict of header meta
    - out_path: output .xlsx path
    """
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Write main table
        df.to_excel(writer, sheet_name='Order', index=False)
        ws = writer.sheets['Order']
        # Freeze header row
        ws.freeze_panes = ws['A2']
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
        # Format numbers (try to detect numeric columns)
        for col_idx, col in enumerate(df.columns, 1):
            if pd.api.types.is_numeric_dtype(df[col]):
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                    for cell in row:
                        cell.number_format = '#,##0.00'
        # Auto-size columns
        for col_idx, col in enumerate(df.columns, 1):
            maxlen = max((len(str(cell.value)) for cell in ws[get_column_letter(col_idx)]), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(maxlen + 2, 40))
        # Add summary sheet
        summary = writer.book.create_sheet('Summary', 0)
        summary['A1'] = 'Order Summary'
        summary['A1'].font = Font(bold=True, size=14)
        for i, (k, v) in enumerate(meta.items(), start=2):
            summary[f'A{i}'] = k.replace('_', ' ').title()
            summary[f'B{i}'] = v
            summary[f'A{i}'].font = Font(bold=True)
            summary[f'B{i}'].alignment = Alignment(horizontal='left')
